"""
Adds Spanish translation columns to dict_new.xlsx:
  pretty_name_es, dataset_es, description_for_ui_es, add_indices_es
Run once from the backend directory, then re-run the ETL processor.
"""

import openpyxl
from pathlib import Path

TRANSLATIONS: dict[str, dict] = {
    "SUR_index_giraudy_2015": {
        "pretty_name_es": "Índice SUR Giraudy (2015)",
        "dataset_es": "Índices de Democracia",
        "description_for_ui_es": "el nivel de democracia subnacional",
        "add_indices_es": " Valores más altos indican mayores niveles de democracia subnacional.",
    },
    "SUR_index_sub_exe": {
        "pretty_name_es": "Índice SUR",
        "dataset_es": "Índices de Democracia",
        "description_for_ui_es": "el nivel de democracia subnacional en el poder ejecutivo",
        "add_indices_es": " Valores más altos indican mayores niveles de democracia subnacional.",
    },
    "contestation_sub_exe": {
        "pretty_name_es": "Contienda Ejecutiva",
        "dataset_es": "Índices de Democracia",
        "description_for_ui_es": "el nivel de contienda subnacional ejecutiva",
        "add_indices_es": " Valores más altos indican mayores niveles de competitividad ejecutiva subnacional.",
    },
    "contestation_sub_leg": {
        "pretty_name_es": "Contienda Legislativa",
        "dataset_es": "Índices de Democracia",
        "description_for_ui_es": "el nivel de contienda subnacional legislativa",
        "add_indices_es": " Valores más altos indican mayores niveles de competitividad legislativa subnacional.",
    },
    "turnover_sub_exe": {
        "pretty_name_es": "Alternancia Ejecutiva",
        "dataset_es": "Índices de Democracia",
        "description_for_ui_es": "el promedio de alternancia del gobernador y del partido",
        "add_indices_es": " Valores más altos indican mayor alternancia ejecutiva subnacional.",
    },
    "clean_elections_sub_exe_giraudy_2015": {
        "pretty_name_es": "Elecciones Limpias Giraudy (2015)",
        "dataset_es": "Índices de Democracia",
        "description_for_ui_es": "el índice de conflicto post-electoral",
        "add_indices_es": " Valores más altos indican elecciones más limpias a nivel ejecutivo subnacional, reflejando menor conflicto post-electoral.",
    },
    "country_name": {
        "pretty_name_es": "País",
        "dataset_es": None,
        "description_for_ui_es": "nombre del país",
        "add_indices_es": None,
    },
    "year": {
        "pretty_name_es": "Año",
        "dataset_es": None,
        "description_for_ui_es": "año calendario",
        "add_indices_es": None,
    },
    "state_name": {
        "pretty_name_es": "Estado",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "nombre de la unidad subnacional",
        "add_indices_es": None,
    },
    "region_name": {
        "pretty_name_es": "Región",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "región geográfica dentro del país",
        "add_indices_es": None,
    },
    "name_head_nat_exe": {
        "pretty_name_es": "Nombre del Presidente",
        "dataset_es": None,
        "description_for_ui_es": "nombre del jefe de estado",
        "add_indices_es": None,
    },
    "sex_head_nat_exe": {
        "pretty_name_es": "Género del Presidente",
        "dataset_es": None,
        "description_for_ui_es": "el género del jefe de estado",
        "add_indices_es": None,
    },
    "term_head_nat_exe": {
        "pretty_name_es": "Mandato del Presidente",
        "dataset_es": "No existe",
        "description_for_ui_es": "fechas de inicio y fin del mandato presidencial",
        "add_indices_es": None,
    },
    "term_length_in_years_nat_exe": {
        "pretty_name_es": "Duración del Mandato Presidencial",
        "dataset_es": None,
        "description_for_ui_es": "duración total del mandato presidencial en años",
        "add_indices_es": None,
    },
    "consecutive_reelection_nat_exe": {
        "pretty_name_es": "Reelección Presidencial",
        "dataset_es": None,
        "description_for_ui_es": "si el presidente fue reelecto",
        "add_indices_es": None,
    },
    "early_exit_nat_exe": {
        "pretty_name_es": "Salida Anticipada del Presidente",
        "dataset_es": None,
        "description_for_ui_es": "si el presidente dejó el cargo antes de completar el mandato",
        "add_indices_es": None,
    },
    "head_party_nat_exe": {
        "pretty_name_es": "Partido del Presidente",
        "dataset_es": None,
        "description_for_ui_es": "partido político del presidente",
        "add_indices_es": None,
    },
    "ideo_party_nat_exe": {
        "pretty_name_es": "Ideología del Partido Presidencial",
        "dataset_es": None,
        "description_for_ui_es": "clasificación ideológica del partido del presidente",
        "add_indices_es": None,
    },
    "year_election_nat_exe": {
        "pretty_name_es": "Año de Elección Presidencial",
        "dataset_es": None,
        "description_for_ui_es": "si se celebraron elecciones presidenciales",
        "add_indices_es": None,
    },
    "name_head_sub_exe": {
        "pretty_name_es": "Nombre del Gobernador",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "nombre del candidato ganador a gobernador",
        "add_indices_es": None,
    },
    "sex_head_sub_exe": {
        "pretty_name_es": "Género del Gobernador",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "el género del gobernador",
        "add_indices_es": None,
    },
    "term_head_sub_exe": {
        "pretty_name_es": "Mandato del Gobernador",
        "dataset_es": "No existe",
        "description_for_ui_es": "fechas de inicio y fin del mandato del gobernador",
        "add_indices_es": None,
    },
    "turnover_head_sub_exe": {
        "pretty_name_es": "Cambio de Gobernador",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "si ocurrió un cambio de gobernador",
        "add_indices_es": None,
    },
    "cumulative_changes_head_sub_exe": {
        "pretty_name_es": "Total de Cambios de Gobernador",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "recuento acumulado de cambios de gobernador",
        "add_indices_es": None,
    },
    "term_length_in_years_sub_exe": {
        "pretty_name_es": "Duración del Mandato del Gobernador",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "duración total del mandato del gobernador en años",
        "add_indices_es": None,
    },
    "early_exit_sub_exe": {
        "pretty_name_es": "Salida Anticipada del Gobernador",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "si el gobernador dejó el cargo antes de tiempo",
        "add_indices_es": None,
    },
    "consecutive_reelection_sub_exe": {
        "pretty_name_es": "Reelección del Gobernador",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "si el gobernador fue reelecto",
        "add_indices_es": None,
    },
    "cumulative_years_in_power_sub_exe": {
        "pretty_name_es": "Años del Gobernador en el Cargo",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "el total de años que el gobernador ha servido en todos sus mandatos",
        "add_indices_es": None,
    },
    "head_party_sub_exe": {
        "pretty_name_es": "Partido del Gobernador",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "el partido político del gobernador",
        "add_indices_es": None,
    },
    "turnover_party_sub_exe": {
        "pretty_name_es": "Cambio de Partido del Gobernador",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "si ocurrió un cambio de partido",
        "add_indices_es": None,
    },
    "cumulative_changes_party_sub_exe": {
        "pretty_name_es": "Total de Cambios de Partido",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "recuento acumulado de cambios de partido",
        "add_indices_es": None,
    },
    "cumulative_years_in_power_party_sub_exe": {
        "pretty_name_es": "Años del Partido en el Poder",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "los años acumulados del partido del gobernador en el poder",
        "add_indices_es": None,
    },
    "ideo_party_sub_exe": {
        "pretty_name_es": "Ideología del Partido del Gobernador",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "la clasificación ideológica del partido del gobernador",
        "add_indices_es": None,
    },
    "electoral_year_sub_exe": {
        "pretty_name_es": "Año de Elección al Gobernador",
        "dataset_es": "No existe",
        "description_for_ui_es": "si se celebraron elecciones al gobernador",
        "add_indices_es": None,
    },
    "enp_sub_exe": {
        "pretty_name_es": "Número Efectivo de Partidos",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "el número efectivo de partidos (Laakso y Taagepera, 1979) en el Ejecutivo",
        "add_indices_es": None,
    },
    "cumulative_electoral_year_sub_exe": {
        "pretty_name_es": "Elecciones Acumuladas Ejecutivas",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "recuento acumulado de elecciones al gobernador",
        "add_indices_es": None,
    },
    "concurrent_with_nat_election_sub_exe": {
        "pretty_name_es": "Elecciones Coincidentes con Nacionales (Eje.)",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "si las elecciones al gobernador y presidenciales fueron concurrentes",
        "add_indices_es": None,
    },
    "alignment_with_nat_sub_exe": {
        "pretty_name_es": "Alineación con el Gobierno Nacional",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "la alineación política entre el presidente y el gobernador",
        "add_indices_es": None,
    },
    "snap_election_sub_exe": {
        "pretty_name_es": "Elección Anticipada",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "si ocurrió una elección anticipada",
        "add_indices_es": None,
    },
    "date_electoral_sub_exe": {
        "pretty_name_es": "Fecha de Elección",
        "dataset_es": "No existe",
        "description_for_ui_es": "fecha de la elección al gobernador",
        "add_indices_es": None,
    },
    "voters_registered_sub_exe": {
        "pretty_name_es": "Votantes Registrados",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "el total de votantes registrados para la elección al gobernador",
        "add_indices_es": None,
    },
    "total_voters_sub_exe": {
        "pretty_name_es": "Total de Votantes",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "el total de personas que votaron en la elección al gobernador",
        "add_indices_es": None,
    },
    "perc_voter_sub_exe": {
        "pretty_name_es": "Participación Electoral %",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "el porcentaje de participación electoral en la elección al gobernador",
        "add_indices_es": None,
    },
    "valid_votes_sub_exe": {
        "pretty_name_es": "Votos Válidos",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "el total de votos válidos en la elección al gobernador",
        "add_indices_es": None,
    },
    "perc_valid_votes_sub_exe": {
        "pretty_name_es": "Votos Válidos %",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "los votos válidos como porcentaje del total de votos",
        "add_indices_es": None,
    },
    "invalid_votes_sub_exe": {
        "pretty_name_es": "Votos Inválidos",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "el total de votos inválidos en la elección al gobernador",
        "add_indices_es": None,
    },
    "perc_invalid_votes_sub_exe": {
        "pretty_name_es": "Votos Inválidos %",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "los votos inválidos como porcentaje del total de votos",
        "add_indices_es": None,
    },
    "votes_candidate_winner_sub_exe": {
        "pretty_name_es": "Votos del Ganador",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "el total de votos del candidato ganador",
        "add_indices_es": None,
    },
    "perc_votes_winner_candidate_sub_exe": {
        "pretty_name_es": "Votos del Ganador %",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "el porcentaje de votos del candidato ganador",
        "add_indices_es": None,
    },
    "second_place_votes_sub_exe": {
        "pretty_name_es": "Votos del Segundo Lugar",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "el total de votos del candidato en segundo lugar",
        "add_indices_es": None,
    },
    "perc_second_place_sub_exe": {
        "pretty_name_es": "Votos del Segundo Lugar %",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "el porcentaje de votos del candidato en segundo lugar",
        "add_indices_es": None,
    },
    "last_place_votes_sub_exe": {
        "pretty_name_es": "Votos del Último Lugar",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "el total de votos del candidato en último lugar",
        "add_indices_es": None,
    },
    "perc_last_place_sub_exe": {
        "pretty_name_es": "Votos del Último Lugar %",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "el porcentaje de votos del candidato en último lugar",
        "add_indices_es": None,
    },
    "margin_victory_sub_exe": {
        "pretty_name_es": "Margen de Victoria",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "la diferencia en porcentaje de votos entre el ganador y el segundo lugar",
        "add_indices_es": None,
    },
    "num_parties_election_contest_sub_exe": {
        "pretty_name_es": "Número de Partidos",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "el número de partidos que compiten en la elección",
        "add_indices_es": None,
    },
    "chamber_election_sub_leg": {
        "pretty_name_es": "Elección por Cámara",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "indica la cámara donde se celebraron las elecciones",
        "add_indices_es": None,
    },
    "enp_sub_leg": {
        "pretty_name_es": "Número Efectivo de Partidos Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "el número efectivo de partidos (Laakso y Taagepera, 1979) en el Legislativo",
        "add_indices_es": None,
    },
    "voters_registered_sub_leg": {
        "pretty_name_es": "Votantes Registrados Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "el total de votantes registrados en elecciones legislativas",
        "add_indices_es": None,
    },
    "total_voters_sub_leg": {
        "pretty_name_es": "Total de Votantes Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "el total de votantes en elecciones legislativas",
        "add_indices_es": None,
    },
    "perc_voters_sub_leg": {
        "pretty_name_es": "Participación Electoral Leg. %",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "la tasa de participación electoral",
        "add_indices_es": None,
    },
    "valid_votes_sub_leg": {
        "pretty_name_es": "Votos Válidos Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "el total de votos válidos",
        "add_indices_es": None,
    },
    "perc_valid_votes_sub_leg": {
        "pretty_name_es": "Votos Válidos Leg. %",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "los votos válidos como % del total",
        "add_indices_es": None,
    },
    "party_name_sub_leg": {
        "pretty_name_es": "Nombre del Partido Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "nombre del partido o coalición",
        "add_indices_es": None,
    },
    "blank_votes_sub_leg": {
        "pretty_name_es": "Votos en Blanco Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "votos en blanco",
        "add_indices_es": None,
    },
    "null_votes_sub_leg": {
        "pretty_name_es": "Votos Nulos Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "votos nulos",
        "add_indices_es": None,
    },
    "challenged_votes_sub_leg": {
        "pretty_name_es": "Votos Impugnados Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "votos impugnados",
        "add_indices_es": None,
    },
    "compensation_for_record_discrepancy_votes_sub_leg": {
        "pretty_name_es": "Votos por Discrepancia de Acta Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "votos ajustados por discrepancia en actas",
        "add_indices_es": None,
    },
    "margin_victory_sub_leg": {
        "pretty_name_es": "Margen de Victoria Leg. %",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "margen entre ganador y segundo lugar",
        "add_indices_es": None,
    },
    "cumulative_elections_year_sub_leg": {
        "pretty_name_es": "Elecciones Acumuladas Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "el recuento acumulado de elecciones legislativas",
        "add_indices_es": None,
    },
    "concurrent_election_with_nat_sub_leg": {
        "pretty_name_es": "Elecciones Leg. Coincidentes con Nacionales",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "si las elecciones legislativas y presidenciales se celebraron el mismo año",
        "add_indices_es": None,
    },
    "num_parties_election_contest_sub_leg": {
        "pretty_name_es": "Número de Partidos Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "los partidos que compiten en la elección",
        "add_indices_es": None,
    },
    "total_seats_in_contest_sub_leg": {
        "pretty_name_es": "Escaños en Disputa Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "el total de escaños en disputa",
        "add_indices_es": None,
    },
    "total_chamber_seats_sub_leg": {
        "pretty_name_es": "Escaños Totales de la Cámara Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "el total de escaños en la legislatura",
        "add_indices_es": None,
    },
    "num_seats_incumbent_sub_leg": {
        "pretty_name_es": "Escaños del Partido Gobernante Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "escaños del partido del gobernador",
        "add_indices_es": None,
    },
    "perc_seats_incumbent_sub_leg": {
        "pretty_name_es": "Escaños del Partido Gobernante Leg. %",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "porcentaje de escaños del partido del gobernador",
        "add_indices_es": None,
    },
    "num_seats_opos_sub_leg": {
        "pretty_name_es": "Escaños de Oposición Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "escaños de los partidos de oposición",
        "add_indices_es": None,
    },
    "perc_seats_opos_sub_leg": {
        "pretty_name_es": "Escaños de Oposición Leg. %",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "porcentaje de escaños de la oposición",
        "add_indices_es": None,
    },
    "chamber_sub_leg": {
        "pretty_name_es": "Tipo de Cámara",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "el número de cámaras legislativas",
        "add_indices_es": None,
    },
    "term_length_in_years_sub_leg": {
        "pretty_name_es": "Duración del Mandato Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "la duración del mandato legislativo en años",
        "add_indices_es": None,
    },
    "renewal_type_sub_leg": {
        "pretty_name_es": "Tipo de Renovación Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "el tipo de renovación de la cámara",
        "add_indices_es": None,
    },
    "electoral_system_sub_leg": {
        "pretty_name_es": "Sistema Electoral Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "el sistema electoral de la cámara",
        "add_indices_es": None,
    },
    "total_seats_party_sub_leg": {
        "pretty_name_es": "Escaños del Partido",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "escaños ganados por un partido",
        "add_indices_es": None,
    },
    "total_votes_party_sub_leg": {
        "pretty_name_es": "Votos del Partido Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "total de votos para un partido o coalición",
        "add_indices_es": None,
    },
    "contested_votes_sub_leg": {
        "pretty_name_es": "Votos en Disputa Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "votos con identidad disputada",
        "add_indices_es": None,
    },
    "date_election_sub_leg": {
        "pretty_name_es": "Fecha de Elección Leg.",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "fecha de las elecciones legislativas subnacionales",
        "add_indices_es": None,
    },
    "perc_votes_party_sub_leg": {
        "pretty_name_es": "Votos del Partido Leg. %",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "porcentaje de votos del partido",
        "add_indices_es": None,
    },
    "perc_seats_party_sub_leg": {
        "pretty_name_es": "Escaños del Partido %",
        "dataset_es": "Elecciones Legislativas",
        "description_for_ui_es": "escaños como % del total en disputa",
        "add_indices_es": None,
    },
    "start_date_head_nat_exe": {
        "pretty_name_es": "Fecha de Inicio (Ejecutivo Nacional)",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "fecha de inicio del mandato del ejecutivo nacional",
        "add_indices_es": None,
    },
    "end_date_head_nat_exe": {
        "pretty_name_es": "Fecha de Fin (Ejecutivo Nacional)",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "fecha de fin del mandato del ejecutivo nacional",
        "add_indices_es": None,
    },
    "cumulative_years_in_power_nat_exe": {
        "pretty_name_es": "Años Acumulados en el Poder (Nacional)",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "años acumulados del jefe de estado nacional en el poder",
        "add_indices_es": None,
    },
    "start_date_head_sub_exe": {
        "pretty_name_es": "Fecha de Inicio (Ejecutivo Subnacional)",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "fecha de inicio del mandato del ejecutivo subnacional",
        "add_indices_es": None,
    },
    "end_date_head_sub_exe": {
        "pretty_name_es": "Fecha de Fin (Ejecutivo Subnacional)",
        "dataset_es": "Ejecutivo",
        "description_for_ui_es": "fecha de fin del mandato del ejecutivo subnacional",
        "add_indices_es": None,
    },
    "election_sub_exe": {
        "pretty_name_es": "Elección Ocurrida",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "si ocurrió una elección ejecutiva ese año (1=sí)",
        "add_indices_es": None,
    },
    "winner_candidate_sub_exe": {
        "pretty_name_es": "Candidato Ganador",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "nombre del candidato ganador a gobernador",
        "add_indices_es": None,
    },
    "date_election_sub_exe": {
        "pretty_name_es": "Fecha de Elección (Ejecutiva)",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "fecha de la elección ejecutiva subnacional",
        "add_indices_es": None,
    },
    "cumulative_elections_year_sub_exe": {
        "pretty_name_es": "Elecciones Acumuladas",
        "dataset_es": "Elecciones Ejecutivas",
        "description_for_ui_es": "recuento acumulado de elecciones ejecutivas subnacionales en un estado",
        "add_indices_es": None,
    },
}

NEW_COLS = ["pretty_name_es", "dataset_es", "description_for_ui_es", "add_indices_es"]


def main() -> None:
    path = Path("data/raw/dict_new.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Read existing headers
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Add new header cells if not already present
    start_col = len(headers) + 1
    existing = set(headers)
    cols_to_add = [c for c in NEW_COLS if c not in existing]

    for i, col_name in enumerate(cols_to_add):
        ws.cell(row=1, column=start_col + i, value=col_name)

    # Refresh headers after adding
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    var_col = col_idx["variable"]
    updated = 0
    missing = []

    for row in ws.iter_rows(min_row=2):
        var_name = row[var_col - 1].value
        if var_name is None:
            continue
        var_name = str(var_name).strip()

        if var_name not in TRANSLATIONS:
            missing.append(var_name)
            continue

        tr = TRANSLATIONS[var_name]
        for col_name in NEW_COLS:
            if col_name in col_idx:
                ws.cell(row=row[0].row, column=col_idx[col_name], value=tr.get(col_name))
        updated += 1

    wb.save(path)
    print(f"Updated {updated} rows. Missing translations for: {missing}")


if __name__ == "__main__":
    main()
