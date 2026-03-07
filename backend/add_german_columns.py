"""
Adds German translation columns to dict_new.xlsx:
  pretty_name_de, dataset_de, description_for_ui_de, add_indices_de
Run once from the backend directory, then re-run the ETL processor.
"""

import openpyxl
from pathlib import Path

TRANSLATIONS: dict[str, dict] = {
    "SUR_index_giraudy_2015": {
        "pretty_name_de": "SUR-Index Giraudy (2015)",
        "dataset_de": "Demokratieindizes",
        "description_for_ui_de": "das Niveau der subnationalen Demokratie",
        "add_indices_de": " Höhere Werte bedeuten höhere Ebenen subnationaler Demokratie.",
    },
    "SUR_index_sub_exe": {
        "pretty_name_de": "SUR-Index",
        "dataset_de": "Demokratieindizes",
        "description_for_ui_de": "das Niveau der subnationalen Demokratie im Exekutivbereich",
        "add_indices_de": " Höhere Werte bedeuten höhere Ebenen subnationaler Demokratie.",
    },
    "contestation_sub_exe": {
        "pretty_name_de": "Exekutiver Wettbewerb",
        "dataset_de": "Demokratieindizes",
        "description_for_ui_de": "das Niveau des subnationalen exekutiven Wettbewerbs",
        "add_indices_de": " Höhere Werte bedeuten größeren exekutiven Parteienwettbewerb.",
    },
    "contestation_sub_leg": {
        "pretty_name_de": "Legislativer Wettbewerb",
        "dataset_de": "Demokratieindizes",
        "description_for_ui_de": "das Niveau des subnationalen legislativen Wettbewerbs",
        "add_indices_de": " Höhere Werte bedeuten größeren legislativen Parteienwettbewerb.",
    },
    "turnover_sub_exe": {
        "pretty_name_de": "Exekutiver Amtswechsel",
        "dataset_de": "Demokratieindizes",
        "description_for_ui_de": "die durchschnittliche Amtswechselrate von Gouverneur und Partei",
        "add_indices_de": " Höhere Werte bedeuten mehr exekutive Amtswechsel.",
    },
    "clean_elections_sub_exe_giraudy_2015": {
        "pretty_name_de": "Saubere Wahlen Giraudy (2015)",
        "dataset_de": "Demokratieindizes",
        "description_for_ui_de": "den Index des postelectoralen Konflikts",
        "add_indices_de": " Höhere Werte bedeuten sauberere subnational-exekutive Wahlen.",
    },
    "country_name": {
        "pretty_name_de": "Land",
        "dataset_de": None,
        "description_for_ui_de": "Name des Landes",
        "add_indices_de": None,
    },
    "year": {
        "pretty_name_de": "Jahr",
        "dataset_de": None,
        "description_for_ui_de": "Kalenderjahr",
        "add_indices_de": None,
    },
    "state_name": {
        "pretty_name_de": "Bundesstaat",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "Name der subnationalen Einheit",
        "add_indices_de": None,
    },
    "region_name": {
        "pretty_name_de": "Region",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "geografische Region im Land",
        "add_indices_de": None,
    },
    "name_head_nat_exe": {
        "pretty_name_de": "Name des Präsidenten",
        "dataset_de": None,
        "description_for_ui_de": "Name des Staatsoberhauptes",
        "add_indices_de": None,
    },
    "sex_head_nat_exe": {
        "pretty_name_de": "Geschlecht des Präsidenten",
        "dataset_de": None,
        "description_for_ui_de": "das Geschlecht des Staatsoberhauptes",
        "add_indices_de": None,
    },
    "term_head_nat_exe": {
        "pretty_name_de": "Amtszeit des Präsidenten",
        "dataset_de": None,
        "description_for_ui_de": "Anfangs- und Enddatum der Amtszeit des Präsidenten",
        "add_indices_de": None,
    },
    "term_length_in_years_nat_exe": {
        "pretty_name_de": "Länge der Präsidialamtszeit",
        "dataset_de": None,
        "description_for_ui_de": "Gesamtdauer der Amtszeit des Präsidenten in Jahren",
        "add_indices_de": None,
    },
    "consecutive_reelection_nat_exe": {
        "pretty_name_de": "Wiederwahl des Präsidenten",
        "dataset_de": None,
        "description_for_ui_de": "ob der Präsident wiedergewählt wurde",
        "add_indices_de": None,
    },
    "early_exit_nat_exe": {
        "pretty_name_de": "Vorzeitiger Rücktritt des Präsidenten",
        "dataset_de": None,
        "description_for_ui_de": "ob der Präsident das Amt vorzeitig verließ",
        "add_indices_de": None,
    },
    "head_party_nat_exe": {
        "pretty_name_de": "Partei des Präsidenten",
        "dataset_de": None,
        "description_for_ui_de": "politische Partei des Präsidenten",
        "add_indices_de": None,
    },
    "ideo_party_nat_exe": {
        "pretty_name_de": "Ideologie der Präsidialpartei",
        "dataset_de": None,
        "description_for_ui_de": "ideologische Einordnung der Partei des Präsidenten",
        "add_indices_de": None,
    },
    "year_election_nat_exe": {
        "pretty_name_de": "Präsidentschaftswahljahr",
        "dataset_de": None,
        "description_for_ui_de": "ob Präsidentschaftswahlen stattfanden",
        "add_indices_de": None,
    },
    "name_head_sub_exe": {
        "pretty_name_de": "Name des Gouverneurs",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "Name des Gouverneurskandidaten",
        "add_indices_de": None,
    },
    "sex_head_sub_exe": {
        "pretty_name_de": "Geschlecht des Gouverneurs",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "das Geschlecht des Gouverneurs",
        "add_indices_de": None,
    },
    "term_head_sub_exe": {
        "pretty_name_de": "Amtszeit des Gouverneurs",
        "dataset_de": None,
        "description_for_ui_de": "Anfangs- und Enddatum der Gouverneursamtszeit",
        "add_indices_de": None,
    },
    "turnover_head_sub_exe": {
        "pretty_name_de": "Gouverneurswechsel",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "ob ein Gouverneurswechsel stattfand",
        "add_indices_de": None,
    },
    "cumulative_changes_head_sub_exe": {
        "pretty_name_de": "Kumulative Gouverneurswechsel",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "kumulierte Anzahl von Gouverneurswechseln",
        "add_indices_de": None,
    },
    "term_length_in_years_sub_exe": {
        "pretty_name_de": "Länge der Gouverneursamtszeit",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "Gesamtdauer der Gouverneursamtszeit in Jahren",
        "add_indices_de": None,
    },
    "early_exit_sub_exe": {
        "pretty_name_de": "Vorzeitiger Rücktritt des Gouverneurs",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "ob der Gouverneur vorzeitig zurücktrat",
        "add_indices_de": None,
    },
    "consecutive_reelection_sub_exe": {
        "pretty_name_de": "Wiederwahl des Gouverneurs",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "ob der Gouverneur wiedergewählt wurde",
        "add_indices_de": None,
    },
    "cumulative_years_in_power_sub_exe": {
        "pretty_name_de": "Kumulative Amtsjahre des Gouverneurs",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "die Gesamtzahl der Jahre des Gouverneurs im Amt",
        "add_indices_de": None,
    },
    "head_party_sub_exe": {
        "pretty_name_de": "Partei des Gouverneurs",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "die politische Partei des Gouverneurs",
        "add_indices_de": None,
    },
    "turnover_party_sub_exe": {
        "pretty_name_de": "Parteiwechsel des Gouverneurs",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "ob ein Parteiwechsel stattfand",
        "add_indices_de": None,
    },
    "cumulative_changes_party_sub_exe": {
        "pretty_name_de": "Kumulative Parteiwechsel",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "kumulierte Anzahl von Parteiwechseln",
        "add_indices_de": None,
    },
    "cumulative_years_in_power_party_sub_exe": {
        "pretty_name_de": "Kumulative Amtsjahre der Partei",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "die kumulierten Jahre der Gouverneurspartei an der Macht",
        "add_indices_de": None,
    },
    "ideo_party_sub_exe": {
        "pretty_name_de": "Ideologie der Gouverneurspartei",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "die ideologische Einordnung der Gouverneurspartei",
        "add_indices_de": None,
    },
    "electoral_year_sub_exe": {
        "pretty_name_de": "Gouverneurswahljahr",
        "dataset_de": None,
        "description_for_ui_de": "ob Gouverneurswahlen stattfanden",
        "add_indices_de": None,
    },
    "enp_sub_exe": {
        "pretty_name_de": "Effektive Anzahl der Parteien",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "die effektive Anzahl der Parteien (Laakso & Taagepera, 1979) im Exekutivbereich",
        "add_indices_de": None,
    },
    "cumulative_electoral_year_sub_exe": {
        "pretty_name_de": "Kumulative Exekutivwahlen",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "kumulierte Anzahl von Gouverneurswahlen",
        "add_indices_de": None,
    },
    "concurrent_with_nat_election_sub_exe": {
        "pretty_name_de": "Gleichzeitige Wahlen (Exekutiv)",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "ob Gouverneurs- und Präsidentschaftswahlen gleichzeitig stattfanden",
        "add_indices_de": None,
    },
    "alignment_with_nat_sub_exe": {
        "pretty_name_de": "Ausrichtung an die Nationalregierung",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "die politische Ausrichtung zwischen Präsident und Gouverneur",
        "add_indices_de": None,
    },
    "snap_election_sub_exe": {
        "pretty_name_de": "Vorzeitige Wahl",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "ob eine vorzeitige Wahl stattfand",
        "add_indices_de": None,
    },
    "date_electoral_sub_exe": {
        "pretty_name_de": "Wahldatum",
        "dataset_de": None,
        "description_for_ui_de": "Datum der Gouverneurswahl",
        "add_indices_de": None,
    },
    "voters_registered_sub_exe": {
        "pretty_name_de": "Eingetragene Wähler",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "Gesamtzahl der eingetragenen Wähler für die Gouverneurswahl",
        "add_indices_de": None,
    },
    "total_voters_sub_exe": {
        "pretty_name_de": "Gesamtzahl der Wähler",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "Gesamtzahl der Personen, die bei der Gouverneurswahl abgestimmt haben",
        "add_indices_de": None,
    },
    "perc_voter_sub_exe": {
        "pretty_name_de": "Wahlbeteiligung %",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "der Prozentsatz der Wahlbeteiligung bei der Gouverneurswahl",
        "add_indices_de": None,
    },
    "valid_votes_sub_exe": {
        "pretty_name_de": "Gültige Stimmen",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "die Gesamtzahl der gültigen Stimmen bei der Gouverneurswahl",
        "add_indices_de": None,
    },
    "perc_valid_votes_sub_exe": {
        "pretty_name_de": "Gültige Stimmen %",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "die gültigen Stimmen als Prozentsatz der Gesamtstimmen",
        "add_indices_de": None,
    },
    "invalid_votes_sub_exe": {
        "pretty_name_de": "Ungültige Stimmen",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "die Gesamtzahl der ungültigen Stimmen bei der Gouverneurswahl",
        "add_indices_de": None,
    },
    "perc_invalid_votes_sub_exe": {
        "pretty_name_de": "Ungültige Stimmen %",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "die ungültigen Stimmen als Prozentsatz der Gesamtstimmen",
        "add_indices_de": None,
    },
    "votes_candidate_winner_sub_exe": {
        "pretty_name_de": "Stimmen des Gewinners",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "die Gesamtstimmen des Gewinners",
        "add_indices_de": None,
    },
    "perc_votes_winner_candidate_sub_exe": {
        "pretty_name_de": "Stimmen des Gewinners %",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "der Stimmanteil des Gewinners",
        "add_indices_de": None,
    },
    "second_place_votes_sub_exe": {
        "pretty_name_de": "Stimmen des Zweitplatzierten",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "die Gesamtstimmen des Zweitplatzierten",
        "add_indices_de": None,
    },
    "perc_second_place_sub_exe": {
        "pretty_name_de": "Stimmen des Zweitplatzierten %",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "der Stimmanteil des Zweitplatzierten",
        "add_indices_de": None,
    },
    "last_place_votes_sub_exe": {
        "pretty_name_de": "Stimmen des Letztplatzierten",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "die Gesamtstimmen des Letztplatzierten",
        "add_indices_de": None,
    },
    "perc_last_place_sub_exe": {
        "pretty_name_de": "Stimmen des Letztplatzierten %",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "der Stimmanteil des Letztplatzierten",
        "add_indices_de": None,
    },
    "margin_victory_sub_exe": {
        "pretty_name_de": "Wahlsieg-Marge",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "die Differenz in Stimmprozenten zwischen Gewinner und Zweitem",
        "add_indices_de": None,
    },
    "num_parties_election_contest_sub_exe": {
        "pretty_name_de": "Anzahl der Parteien",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "die Anzahl der an der Wahl teilnehmenden Parteien",
        "add_indices_de": None,
    },
    "chamber_election_sub_leg": {
        "pretty_name_de": "Wahl nach Kammer",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "gibt an, in welcher Kammer die Wahlen stattfanden",
        "add_indices_de": None,
    },
    "enp_sub_leg": {
        "pretty_name_de": "Effektive Anzahl der Parteien Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "die effektive Anzahl der Parteien (Laakso & Taagepera, 1979) im Legislativbereich",
        "add_indices_de": None,
    },
    "voters_registered_sub_leg": {
        "pretty_name_de": "Eingetragene Wähler Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "die Gesamtzahl der eingetragenen Wähler bei Legislativwahlen",
        "add_indices_de": None,
    },
    "total_voters_sub_leg": {
        "pretty_name_de": "Gesamtzahl der Wähler Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "die Gesamtzahl der Wähler bei Legislativwahlen",
        "add_indices_de": None,
    },
    "perc_voters_sub_leg": {
        "pretty_name_de": "Wahlbeteiligung Leg. %",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "die Wahlbeteiligungsrate",
        "add_indices_de": None,
    },
    "valid_votes_sub_leg": {
        "pretty_name_de": "Gültige Stimmen Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "die Gesamtzahl der gültigen Stimmen",
        "add_indices_de": None,
    },
    "perc_valid_votes_sub_leg": {
        "pretty_name_de": "Gültige Stimmen Leg. %",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "die gültigen Stimmen als Prozentsatz",
        "add_indices_de": None,
    },
    "party_name_sub_leg": {
        "pretty_name_de": "Name der Partei Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "Name der Partei oder Koalition",
        "add_indices_de": None,
    },
    "blank_votes_sub_leg": {
        "pretty_name_de": "Leere Stimmen Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "leere Stimmen",
        "add_indices_de": None,
    },
    "null_votes_sub_leg": {
        "pretty_name_de": "Nichtige Stimmen Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "nichtige Stimmen",
        "add_indices_de": None,
    },
    "challenged_votes_sub_leg": {
        "pretty_name_de": "Angefochtene Stimmen Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "angefochtene Stimmen",
        "add_indices_de": None,
    },
    "compensation_for_record_discrepancy_votes_sub_leg": {
        "pretty_name_de": "Stimmen für Protokolldiskrepanz Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "Stimmen korrigiert wegen Protokolldiskrepanz",
        "add_indices_de": None,
    },
    "margin_victory_sub_leg": {
        "pretty_name_de": "Wahlsieg-Marge Leg. %",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "Marge zwischen Gewinner und Zweitem",
        "add_indices_de": None,
    },
    "cumulative_elections_year_sub_leg": {
        "pretty_name_de": "Kumulative Legislativwahlen",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "kumulierte Anzahl von Legislativwahlen",
        "add_indices_de": None,
    },
    "concurrent_election_with_nat_sub_leg": {
        "pretty_name_de": "Gleichzeitige Leg.-Wahl mit Nationalwahl",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "ob Legislativ- und Präsidentschaftswahlen im selben Jahr stattfanden",
        "add_indices_de": None,
    },
    "num_parties_election_contest_sub_leg": {
        "pretty_name_de": "Anzahl der Parteien Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "die an der Wahl teilnehmenden Parteien",
        "add_indices_de": None,
    },
    "total_seats_in_contest_sub_leg": {
        "pretty_name_de": "Sitze im Wettbewerb Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "die Gesamtzahl der zu besetzenden Sitze",
        "add_indices_de": None,
    },
    "total_chamber_seats_sub_leg": {
        "pretty_name_de": "Gesamtsitze der Kammer Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "die Gesamtzahl der Sitze in der Legislative",
        "add_indices_de": None,
    },
    "num_seats_incumbent_sub_leg": {
        "pretty_name_de": "Sitze der Regierungspartei Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "Sitze der Gouverneurspartei",
        "add_indices_de": None,
    },
    "perc_seats_incumbent_sub_leg": {
        "pretty_name_de": "Sitze der Regierungspartei Leg. %",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "Prozentsatz der Sitze der Gouverneurspartei",
        "add_indices_de": None,
    },
    "num_seats_opos_sub_leg": {
        "pretty_name_de": "Oppositionssitze Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "Sitze der Oppositionsparteien",
        "add_indices_de": None,
    },
    "perc_seats_opos_sub_leg": {
        "pretty_name_de": "Oppositionssitze Leg. %",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "Prozentsatz der Oppositionssitze",
        "add_indices_de": None,
    },
    "chamber_sub_leg": {
        "pretty_name_de": "Kammertyp",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "Anzahl der Gesetzgebungskammern",
        "add_indices_de": None,
    },
    "term_length_in_years_sub_leg": {
        "pretty_name_de": "Amtsdauer Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "Dauer des legislativen Mandats in Jahren",
        "add_indices_de": None,
    },
    "renewal_type_sub_leg": {
        "pretty_name_de": "Erneuerungstyp Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "die Art der Kammererneuerung",
        "add_indices_de": None,
    },
    "electoral_system_sub_leg": {
        "pretty_name_de": "Wahlsystem Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "das Wahlsystem der Kammer",
        "add_indices_de": None,
    },
    "total_seats_party_sub_leg": {
        "pretty_name_de": "Sitze der Partei",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "von einer Partei gewonnene Sitze",
        "add_indices_de": None,
    },
    "total_votes_party_sub_leg": {
        "pretty_name_de": "Stimmen der Partei Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "Gesamtstimmen für eine Partei oder Koalition",
        "add_indices_de": None,
    },
    "contested_votes_sub_leg": {
        "pretty_name_de": "Streitige Stimmen Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "Stimmen mit strittiger Zuordnung",
        "add_indices_de": None,
    },
    "date_election_sub_leg": {
        "pretty_name_de": "Wahldatum Leg.",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "Datum der subnationalen Legislativwahlen",
        "add_indices_de": None,
    },
    "perc_votes_party_sub_leg": {
        "pretty_name_de": "Stimmen der Partei Leg. %",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "Stimmanteil der Partei",
        "add_indices_de": None,
    },
    "perc_seats_party_sub_leg": {
        "pretty_name_de": "Sitze der Partei %",
        "dataset_de": "Legislativwahlen",
        "description_for_ui_de": "Sitze als Prozentsatz der Gesamtsitze",
        "add_indices_de": None,
    },
    "start_date_head_nat_exe": {
        "pretty_name_de": "Startdatum (Nationaler Exekutive)",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "Startdatum des Mandats des nationalen Exekutivchefs",
        "add_indices_de": None,
    },
    "end_date_head_nat_exe": {
        "pretty_name_de": "Enddatum (Nationaler Exekutive)",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "Enddatum des Mandats des nationalen Exekutivchefs",
        "add_indices_de": None,
    },
    "cumulative_years_in_power_nat_exe": {
        "pretty_name_de": "Kumulative Amtsjahre (National)",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "kumulierte Amtsjahre des nationalen Staatsoberhauptes",
        "add_indices_de": None,
    },
    "start_date_head_sub_exe": {
        "pretty_name_de": "Startdatum (Subnationaler Exekutive)",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "Startdatum des Mandats des subnationalen Exekutivchefs",
        "add_indices_de": None,
    },
    "end_date_head_sub_exe": {
        "pretty_name_de": "Enddatum (Subnationaler Exekutive)",
        "dataset_de": "Exekutive",
        "description_for_ui_de": "Enddatum des Mandats des subnationalen Exekutivchefs",
        "add_indices_de": None,
    },
    "election_sub_exe": {
        "pretty_name_de": "Wahl stattgefunden",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "ob eine Exekutivwahl in diesem Jahr stattfand (1=ja)",
        "add_indices_de": None,
    },
    "winner_candidate_sub_exe": {
        "pretty_name_de": "Gewinner-Kandidat",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "Name des Gouverneurs-Gewinners",
        "add_indices_de": None,
    },
    "date_election_sub_exe": {
        "pretty_name_de": "Wahldatum (Exekutiv)",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "Datum der subnationalen Exekutivwahl",
        "add_indices_de": None,
    },
    "cumulative_elections_year_sub_exe": {
        "pretty_name_de": "Kumulative Wahlen",
        "dataset_de": "Exekutivwahlen",
        "description_for_ui_de": "kumulierte Anzahl subnationaler Exekutivwahlen in einem Staat",
        "add_indices_de": None,
    },
}

NEW_COLS = ["pretty_name_de", "dataset_de", "description_for_ui_de", "add_indices_de"]


def main() -> None:
    path = Path("data/raw/dict_new.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    start_col = len(headers) + 1
    existing = set(headers)
    cols_to_add = [c for c in NEW_COLS if c not in existing]

    for i, col_name in enumerate(cols_to_add):
        ws.cell(row=1, column=start_col + i, value=col_name)

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    var_col = col_idx["variable"]
    updated = 0
    missing = []

    for row in ws.iter_rows(min_row=2):
        var_name = row[var_col - 1].value
        if var_name is None:
            continue
        var_name = str(var_name).strip()

        if var_name not in TRANSLATIONS:
            missing.append(var_name)
            continue

        tr = TRANSLATIONS[var_name]
        for col_name in NEW_COLS:
            if col_name in col_idx:
                ws.cell(row=row[0].row, column=col_idx[col_name], value=tr.get(col_name))
        updated += 1

    wb.save(path)
    print(f"Updated {updated} rows. Missing translations for: {missing}")


if __name__ == "__main__":
    main()
